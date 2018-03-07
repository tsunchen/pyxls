#!/usr/bin/env python
# -*- coding:utf-8 -*-

"""
@author: tsunc & zadmine
@software: PyCharm ommunity Edition
@time: 2018/3/7 10:10
"""
'''
（1）用load_workbook函数打开excel文件，返回一个工作簿对象

（2）用工作簿对象获取所有的sheet

（3）第一个for循环遍历所有sheet

（4）每次遍历sheet时，先用get_sheet_by_name获取要遍历的sheet

（5）开始下一个for循环，循环遍历这个sheet中的所有行

（6）在遍历每一行的for循环中，遍历每一列，把每一列的值，通过join函数拼接起来，这里用ljust函数进行左对齐


'''



import openpyxl

import sys
reload(sys)
sys.setdefaultencoding('utf-8')


lines = []
keys = []


def readfile(filename):  
    with open(filename,'r') as f:  
        for line in f.readlines():  
            linestr = line.strip()  
            print linestr  
            #print linestr.split(".")[:3]
            linestrlist2 = linestr.split(".")[:3][0] + "." + linestr.split(".")[:3][1] + "." + linestr.split(".")[:3][2]
            print linestrlist2
            ##linestrlist = linestr.split("\t")  
            #print linestrlist 
            ##keys.append(str(linestr.encode('utf-8')))
            keys.append(str(linestrlist2.encode('utf-8')))
            ##linelist = map(int,linestrlist)# 方法一  
            # linelist = [int(i) for i in linestrlist] # 方法二  
            ##print linelist  

#wb = openpyxl.load_workbook('E:/pythonlab/ITSM4.xlsx')
#wb = openpyxl.load_workbook('E:/pythonlab/ba-20170516.xlsx')  
wb = openpyxl.load_workbook('E:/pythonlab/ITSM4.xlsx') 
readfile("E:/pythonlab/keys_file.txt")

#获取workbook中所有的表格  
sheets = wb.get_sheet_names()  
  
print(sheets)  


#循环遍历所有sheet  
for i in range(len(sheets)):  
    sheet= wb.get_sheet_by_name(sheets[i])  
      
    #print('\n\nNumber '+str(i+1)+' sheet: ' + sheet.title+'->>>')  
    print('\n\n第'+str(i+1)+'个sheet: ' + sheet.title+'->>>')  
  
    for r in range(1,sheet.max_row+1):  
        if r == 1:
            lines.append('\n'+''.join([str(sheet.cell(row=r,column=c).value).ljust(17) for c in range(1,sheet.max_column+1)] ))  
            #print('\n'+''.join([str(sheet.cell(row=r,column=c).value).ljust(17) for c in range(1,sheet.max_column+1)] ))  
        else:  
        	lines.append(''.join([str(sheet.cell(row=r,column=c).value).ljust(20) for c in range(1,sheet.max_column+1)] ))
            #print(''.join([str(sheet.cell(row=r,column=c).value).ljust(20) for c in range(1,sheet.max_column+1)] ))  
    print len(lines)


print (len(lines))


'''

'''

#keys = [u"211.152.61.4", u"211.152.61.5",u"固定"]
print keys




#f = open("E:/pythonlab/keys_file0.txt", "r")
#g = f.read()
#for k in g:
#	keys.append(str(k))
#	print k
#f.close()

print keys

'''
for i in range(len(lines)):
	if lines[i].find(u"固定") == -1:
	  print "No '0' here!"
	else:
	  print "Found '1' in the lines."
	  print lines[i]
'''

keysdict = {}
for item in keys:
	if item in keysdict:
		keysdict[item.encode('utf-8')] += 1
	else:
		keysdict[item.encode('utf-8')] = 0

for (d,x) in keysdict.items():
	print "key:"+d+", count:"+str(x)

for i in range(len(lines)):
    for j in range(len(keys)):
    	if lines[i].find(keys[j].encode('utf-8')) != -1:
    		keysdict[keys[j].encode('utf-8')] += 1

print ("Result of Line Query Count: >>>>>>")

for (d,x) in keysdict.items():
	print "      key:"+d+", count:"+str(x)

######
keysdictdetail = {}
for item in keys:
	if item in keysdictdetail:
		keysdictdetail[item.encode('utf-8')] += ''
	else:
		keysdictdetail[item.encode('utf-8')] = []

for (d,x) in keysdictdetail.items():
	print "key:"+d+", detail:"
	for n in range(len(x)):
		print x[n]

for i in range(len(lines)):
    for j in range(len(keys)):
    	if lines[i].find(keys[j].encode('utf-8')) != -1:
    		keysdictdetail[keys[j].encode('utf-8')].append(str(lines[i]))
    	#else:
    		#keysdictdetail[keys[j].encode('utf-8')] += ''
    		#keysdict[keys[j].encode('utf-8')] += 1

print ("Result of Line Query Detail: >>>>>>")

for (d,x) in keysdictdetail.items():
	print "      key:"+d+", detail:"
	for n in range(len(x)):
		print "::::::", x[n]
